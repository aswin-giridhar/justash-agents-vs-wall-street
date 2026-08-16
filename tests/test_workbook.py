import pytest
from openpyxl import load_workbook

from avws.workbook import write_workbook

HD_VALUES = {
    "Net sales": 45123.4,
    "Adjusted diluted EPS": 4.6789,
    "Comparable sales, total company": 1.2345,
}


def test_writes_all_three_values_and_keeps_summary_sheet():
    path = write_workbook("HD", HD_VALUES)
    workbook = load_workbook(path)
    assert "Summary" in workbook.sheetnames
    written = [c.value for row in workbook["Summary"].iter_rows() for c in row]
    assert 45123.4 in written
    assert 4.68 in written  # EPS rounded to reported precision
    assert 1.23 in written  # percentage points, not a fraction


def test_values_land_on_the_row_matching_their_label():
    path = write_workbook("ADI", {
        "Revenue": 3958.0,
        "Adjusted diluted EPS": 3.35,
        "Adjusted gross margin": 73.2,
    })
    sheet = load_workbook(path)["Summary"]
    rows = {sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=3).value
            for r in range(1, sheet.max_row + 1)}
    assert rows["Revenue"] == 3958.0
    assert rows["Adjusted gross margin"] == 73.2


def test_refuses_to_write_a_missing_metric():
    with pytest.raises(ValueError, match="missing"):
        write_workbook("HD", {"Net sales": 45000.0})


def test_refuses_a_nan_value():
    with pytest.raises(ValueError, match="non-finite"):
        write_workbook("HD", {**HD_VALUES, "Net sales": float("nan")})


def test_unknown_ticker_raises():
    with pytest.raises(KeyError):
        write_workbook("XXX", {})
