from openpyxl import load_workbook
from pathlib import Path
p = Path(r"E:\Hackathon\agents_vs_wallstreet_hackathon\starter\submission_openstocks\HAS-2026H2-forecast-template.xlsx")
wb = load_workbook(p); ws = wb["Summary"]
# H2 25 = 0.4764. H1 26 was -8.6% YoY; Q4 26 improved to -5% (Q4 trading update),
# so H2 26 at about -7% YoY is consistent with both, where -16.9% is not.
fixes = {
    "Net Fees": (0.4431, "0.4764 H2 25 x (1 - 7.0%), the midpoint of the -8.6% H1 run-rate and the -5% Q4 exit rate"),
    "Net Fees - Germany": (0.1442, "0.1518 H2 25 x (1 - 5.0%), the Q4 Germany actual decline"),
}
for r in range(2, ws.max_row+1):
    lbl = ws.cell(row=r, column=1).value
    if lbl in fixes:
        old = ws.cell(row=r, column=6).value
        new, why = fixes[lbl]
        ws.cell(row=r, column=6, value=new)
        print(f"  {lbl:<34} {old} -> {new}   {why}")
    elif lbl:
        print(f"  {lbl:<34} {ws.cell(row=r,column=6).value}  (unchanged)")
wb.save(p)
print("\nHAS workbook updated")
