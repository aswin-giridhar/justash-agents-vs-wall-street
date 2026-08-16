"""Write completed OpenStocks workbooks from the supplied templates.

Template layout, verified in Task 0 and identical across all four files: sheet
`Summary`, header row 6, metrics on rows 7-9 with the label in column A, units in
column B and the empty forecast cell in column C. We locate rows by matching the
label text rather than by hardcoding row numbers, so a template revision cannot
silently move a value into the wrong metric.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from avws.config import SUBMISSION_DIR, TEMPLATE_DIR
from avws.registry import Metric, metrics_for

SHEET = "Summary"
LABEL_COL = 1
FORECAST_COL = 3


def _find_label_row(worksheet, label: str) -> int:
    for row in range(1, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=LABEL_COL).value
        if isinstance(cell, str) and cell.strip() == label:
            return row
    raise ValueError(f"label not found in template {SHEET!r} sheet: {label!r}")


def _round_for(metric: Metric, value: float) -> float:
    """Round to the precision the metric is reported at.

    Percentages and EPS are reported to two decimals; money figures in millions to
    whole units. Rounding beyond the reported precision implies a confidence the
    forecast does not have.
    """
    if metric.is_percentage or metric.is_eps:
        return round(float(value), 2)
    return round(float(value), 1)


def write_workbook(ticker: str, values: dict[str, float]) -> Path:
    """Write one company's workbook. `values` maps metric label to number.

    Raises if any of the company's three metrics is missing or non-finite. A blank
    cell scores 5.0 under the competition rules, so refusing to write an incomplete
    workbook is safer than emitting one.
    """
    metrics = metrics_for(ticker)
    if not metrics:
        raise KeyError(f"unknown ticker: {ticker}")

    missing = [m.label for m in metrics if m.label not in values]
    if missing:
        raise ValueError(f"missing metric values for {ticker}: {missing}")

    template = TEMPLATE_DIR / metrics[0].output_file
    workbook = load_workbook(template)
    worksheet = workbook[SHEET]

    for metric in metrics:
        value = values[metric.label]
        if value is None or value != value or value in (float("inf"), float("-inf")):
            raise ValueError(f"non-finite value for {metric.key}: {value!r}")
        row = _find_label_row(worksheet, metric.label)
        worksheet.cell(row=row, column=FORECAST_COL, value=_round_for(metric, value))

    SUBMISSION_DIR.mkdir(parents=True, exist_ok=True)
    out = SUBMISSION_DIR / metrics[0].output_file
    workbook.save(out)
    return out
