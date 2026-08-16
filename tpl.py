from openpyxl import load_workbook
import glob, os
for p in sorted(glob.glob(r"E:\Hackathon\agents_vs_wallstreet_hackathon\Context\*forecast-template.xlsx")):
    wb = load_workbook(p)
    print("="*70); print(os.path.basename(p), "sheets:", wb.sheetnames)
    for ws in wb.worksheets:
        print(f"-- {ws.title}  {ws.max_row}x{ws.max_column}")
        for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row,30)):
            cells=[(c.coordinate,c.value) for c in r if c.value is not None]
            if cells: print("   ", cells)
        break
