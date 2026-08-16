from openpyxl import load_workbook
import glob, os
for p in sorted(glob.glob(r"E:\Hackathon\agents_vs_wallstreet_hackathon\starter\submission_openstocks\*.xlsx")):
    ws = load_workbook(p)["Summary"]
    print(f"\n{os.path.basename(p)}   [{ws['F1'].value}]")
    for r in range(2, ws.max_row+1):
        print(f"   {ws.cell(row=r,column=1).value:<36} {ws.cell(row=r,column=6).value:>10}  {ws.cell(row=r,column=2).value}")
