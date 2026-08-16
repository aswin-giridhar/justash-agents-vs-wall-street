"""Fill the OpenStocks upload templates from our forecasts.

The upload templates differ from the starter templates: units are BILLIONS, the
periods differ (HD is 2027Q2, Hays is 2026H2 not the full year), and three metrics
are ones the system did not forecast. Provenance for every cell is printed so the
hand-derived ones are visible rather than implied to be system output.
"""

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

CTX = Path(r"E:\Hackathon\agents_vs_wallstreet_hackathon\Context")
OUT = Path(r"E:\Hackathon\agents_vs_wallstreet_hackathon\starter\submission_openstocks")
OUT.mkdir(exist_ok=True)

f = json.loads(Path("forecasts.json").read_text(encoding="utf-8"))
V = {c["ticker"]: c["values"] for c in f["companies"]}


def g(t, k, default=None):
    return V.get(t, {}).get(k, default)


# --- values, in the units each template asks for -----------------------------
# SYSTEM = produced by the pipeline. DERIVED = arithmetic on system output.
# EXTRAPOLATED = not forecast by the system; trend of the three printed periods.
PLAN = {
    "HD-2027Q2-forecast-template.xlsx": {
        "Net sales":  (g("HD", "Net sales", 47502) / 1000, "SYSTEM  USDm->bn"),
        "Diluted EPS": (round(g("HD", "Adjusted diluted EPS", 4.76) - 0.13, 3),
                        "DERIVED  our adjusted EPS less the recent GAAP-vs-adjusted gap"),
        "Comparable sales": (round(g("HD", "Comparable sales, total company", 1.0), 2),
                             "SYSTEM  percentage points"),
    },
    "ADI.O-2026Q3-forecast-template.xlsx": {
        "Revenue": (g("ADI", "Revenue", 3899) / 1000, "SYSTEM  USDm->bn"),
        "Adjusted Diluted EPS": (round(g("ADI", "Adjusted diluted EPS", 3.67), 3),
                                 "SYSTEM"),
        # Industrial ran 1.422 -> 1.4893 -> 1.7994 and is ADI's largest end market;
        # growing with total revenue at its recent share.
        "Industrial Revenue": (1.90, "EXTRAPOLATED  not forecast by the system"),
    },
    "HAS-2026H2-forecast-template.xlsx": {
        # The template wants the SECOND HALF, our forecast is the full year.
        "Net Fees": (round(g("HAS", "Net fees", 849) / 1000 - 0.4533, 4),
                     "DERIVED  our FY forecast less reported H1 26 of 0.4533"),
        "Pre-exceptional Operating Profit": (
            round(g("HAS", "Pre-exceptional operating profit", 45.7) / 1000 - 0.0201, 4),
            "DERIVED  our FY forecast less reported H1 26 of 0.0201"),
        "Net Fees - Germany": (0.1400,
                               "EXTRAPOLATED  0.1571 -> 0.1518 -> 0.1455 trend"),
    },
    "DE-2026Q3-forecast-template.xlsx": {
        "Net Sales": (g("DE", "Worldwide net sales and revenues", 10615) / 1000,
                      "SYSTEM  USDm->bn"),
        "Diluted EPS": (round(g("DE", "Diluted EPS (GAAP)", 4.66), 3), "SYSTEM"),
        # We forecast the segment's operating PROFIT, not its net sales.
        "Production & Precision Ag Net Sales": (
            4.45, "EXTRAPOLATED  4.74 -> 3.163 -> 4.503, Q3 seasonally similar to Q2"),
    },
}

for name, rows in PLAN.items():
    src = CTX / name
    dst = OUT / name
    shutil.copy(src, dst)
    wb = load_workbook(dst)
    ws = wb["Summary"]
    print(f"\n=== {name}  (forecast column {ws['F1'].value}) ===")
    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label in rows:
            value, note = rows[label]
            ws.cell(row=r, column=6, value=round(float(value), 4))
            print(f"   {label:<38} {value:>10.4f}  {ws.cell(row=r,column=2).value:<10} {note}")
        elif label:
            print(f"   {label:<38} {'!! UNFILLED':>10}")
    wb.save(dst)

print(f"\nwrote {len(PLAN)} workbooks to {OUT}")
