from openpyxl import load_workbook
from pathlib import Path
D = Path(r"E:\Hackathon\agents_vs_wallstreet_hackathon\starter\submission_openstocks")

# Both were extrapolations. Each is a SHARE of a total we already forecast, and the
# template prints three quarters of both the segment and the total, so the share is
# computable and the value becomes arithmetic on system output.
jobs = [
  ("ADI.O-2026Q3-forecast-template.xlsx", "Industrial Revenue", "Revenue",
   [(1.422,3.0761),(1.4893,3.1603),(1.7994,3.6235)], "rising", 0.505),
  ("DE-2026Q3-forecast-template.xlsx", "Production & Precision Ag Net Sales", "Net Sales",
   [(4.74,10.579),(3.163,8.001),(4.503,11.778)], "mean", None),
]
for fn, seg_label, tot_label, hist, mode, override in jobs:
    p = D/fn; wb = load_workbook(p); ws = wb["Summary"]
    rows = {ws.cell(row=r,column=1).value: r for r in range(2, ws.max_row+1)}
    total = ws.cell(row=rows[tot_label], column=6).value
    shares = [s/t for s,t in hist]
    share = override if override else sum(shares)/len(shares)
    new = round(total*share, 4)
    old = ws.cell(row=rows[seg_label], column=6).value
    ws.cell(row=rows[seg_label], column=6, value=new)
    print(f"{fn}")
    print(f"   shares: {', '.join(f'{s:.1%}' for s in shares)}  -> using {share:.1%} ({mode})")
    print(f"   {seg_label}: {old} -> {new}   = our {tot_label} forecast {total} x {share:.3f}")
    wb.save(p)
